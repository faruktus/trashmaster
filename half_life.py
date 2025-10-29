import json
import locale
locale.setlocale(locale.LC_ALL, 'de_DE.utf8')
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from itertools import count
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from calendar import Calendar

try:
    with open("deimuada.json", mode="r", encoding="utf-8") as read_file:
           dict_interval = json.load(read_file)
except:
    print("no file")

starting_date = date.today()
plusyear = starting_date + relativedelta(months=11)

# Compare 2 Lists for duplicates
def f_CompList(firstList, secondList):
    wholeList = len(firstList) + len(secondList)
    uniqueList = len(set(firstList + secondList))
    return wholeList - uniqueList

# increase every item in list for +1 day everytime function is called (max + 6 days)
# e.g. list = [3.1.25, 4.1.25] 1st call: [4.1.25, 5.1.25] 2nd call: [5.1.25, 6.1.25]
cnt = count(1)
def f_IterList(listi):
    global cnt
    nextday = next(cnt) % 6
    return [x + relativedelta(days=nextday) for x in listi]


# compare 2 datelists - raise v_iterlist by 1 day while 0 duplicates 
#                       OR return list with least duplicates
def f_CompMultilist(v_iterlist, list2):
    templist = []
    for x in range(6):
        iterlist = f_IterList(v_iterlist)
        complist = f_CompList(iterlist, list2)
        if complist == 0:
            return iterlist
        else:
            templist.append([complist, iterlist])
    return sorted(templist)[0][1]

# dict_days: BEREICH = [Liste von errechneten Tagen im Format datetime.date für 1 Jahr]
dict_days={}
try:
    global calc
    for k,v in dict_interval.items():
        if v[0] == "Woche":
          calc = int(334/47/v[1])
        if v[0]  == "Monat":
          calc = int(334/11/v[1])
        if v[0]  == "Jahr":
          calc = int(334/v[1])

        date_list=[]
        new_start = starting_date
        while starting_date <= (plusyear-relativedelta(days=calc)):
            starting_date += relativedelta(days=calc)
            #print(starting_date)
            date_list.append(starting_date)
            #listi.append(starting_date.strftime("%d, %m, %Y"))
            dict_days[k] = date_list
        #print()

        # spread multiple tasks on same day to different days
        dict_values = [y for x in dict_days.values() for y in x]
        dict_days[k] = f_CompMultilist(date_list, dict_values)

        starting_date=new_start   
except:
    print("not enough/wrong data")


# >>>>>>>>>>>>>>>>> WRITE DATA TO EXCEL <<<<<<<<<<<<<<<<


wb = Workbook()
sheet = wb.active

sheet.title = "Data"

starting_year = int(starting_date.strftime("%Y"))
starting_month = int(starting_date.strftime("%m"))

monthnames = []
for x in range(1,13):
    monthnames.append(starting_date.strftime("%B %y"))
    starting_date += relativedelta(months=1)

# e.g.: dict_month = {'Oktober 25': [1,2...,31]}
obj=Calendar()
dict_month={}
for index, x in enumerate(monthnames):
    dict_month[x] = [x for x in obj.itermonthdays(starting_year, 
                                                 ((index + starting_month-1)%12)+1) if x]
    
# write dict_month to excel sheet 
col = [x for x in range(1, 24, 2)]
for i, (k, v) in enumerate(dict_month.items()):
    sheet.cell(row=1, column=col[i], value=k)
    # Format cells and merge >>>
    sheet.cell(row=1, column=col[i]).font = Font(name='Calibri', size=12, bold=True)
    sheet.cell(row=1, column=col[i]).alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_column = col[i], end_column = col[i]+1, start_row=1, end_row=1)
    # <<< Format cells and merge
    row = 2
    for day in v:
        sheet.cell(row=row, column=col[i], value=day)
        # Format >>>
        sheet.cell(row=row, column=col[i]).font = Font(name='Calibri', size=12, bold=True)
        sheet.cell(row=row, column=col[i]).alignment = Alignment(horizontal='center', 
                                                        vertical='center')
        # <<< Format
        row += 1

# getCord = Function to get cell from a list formated ['Oktober 25', '31']
header = sheet['A1':'X1'][0] 
titlecol = sheet['A']
def getCord(listi):
    global header, titlecol
    column = [x for x in header if x.value==listi[0]][0].column 
    row = [x for x in titlecol if x.value==listi[1]][0].row
    return sheet.cell(column=column+1, row=row)
    
# modify dict_days values to a list with excel coordinates e.g. {'Garderobe': [<Cell 'Data'.D9>,..]}
for k, v in dict_days.items():
    temp_list=[]
    for values in v:
        temp_list.append(getCord([values.strftime("%B %y"), int(values.strftime("%d"))]))
    dict_days[k] = temp_list

# write areas to the correct date in excel
for k,v in dict_days.items():
    for val in v:
        # format
        val.font = Font(name='Calibri', size=12)
        if not val.value:
            val.value = k[:3]
        else:
            val.value = val.value + " | " + str(k)[:3]

# set width of columns
temp_list=[]
for index, col_number in enumerate(range(2,25,2)):
    # adjust width for columns with daynumbers
    sheet.column_dimensions[get_column_letter(col_number-1)].width=4
    # adjust width for columns with areas
    col = sheet.iter_cols(min_col=col_number, max_col=col_number, min_row=2, values_only=True)
    col = sorted([len(str(x)) for y in col for x in y if x])
    # check if string of header or daytasks is bigger
    header_width = len(monthnames[index])
    if col:
        max_width = max([col[-1], header_width]) + 1
        sheet.column_dimensions[get_column_letter(col_number)].width=max_width


wb.save("half_life.xlsx")

