import openpyxl
import locale
locale.setlocale(locale.LC_ALL, 'de_DE.utf8')
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import calendar
from calendar import Calendar, TextCalendar
from newjason import dict_days, starting_date
from dateutil.relativedelta import relativedelta

wb = Workbook()
sheet = wb.active

sheet.title = "Data"

starting_year = int(starting_date.strftime("%Y"))
starting_month = int(starting_date.strftime("%m"))

monthnames = []
for x in range(1,13):
    monthnames.append(starting_date.strftime("%B %y"))
    starting_date += relativedelta(months=1)

# e.g.: dict_month = {'Oktober 25': [1,2...,31]}
obj=Calendar()
dict_month={}
for index, x in enumerate(monthnames):
    dict_month[x] = [x for x in obj.itermonthdays(starting_year, 
                                                 ((index + starting_month-1)%12)+1) if x]
    
# write dict_month to excel sheet and merge columns in header
col = [x for x in range(1, 24, 2)]
for i, (k, v) in enumerate(dict_month.items()):
    sheet.cell(row=1, column=col[i], value=k)
    sheet.merge_cells(start_column = col[i], end_column = col[i]+1, start_row=1, end_row=1)
    row = 2
    for day in v:
        sheet.cell(row=row, column=col[i], value=day)
        row += 1

# getCord = Function to get cell from a list formated ['Oktober 25', 
header = sheet['A1':'X1'][0] 
titlecol = sheet['A']
def getCord(listi):
    global header, titlecol
    column = [x for x in header if x.value==listi[0]][0].column 
    row = [x for x in titlecol if x.value==listi[1]][0].row
    return sheet.cell(column=column+1, row=row)
    
for k, v in dict_days.items():
    temp_list=[]
    for values in v:
        temp_list.append(getCord([values.strftime("%B %y"), int(values.strftime("%d"))]))
    dict_days[k] = temp_list

for k,v in dict_days.items():
    for val in v:
        if not val.value:
            val.value = k[:3]
        else:
            val.value = val.value + " | " + str(k)[:3]

temp_list=[]
for col_number in range(2,25,2):
    # adjust width for columns with daynumbers
    sheet.column_dimensions[get_column_letter(col_number-1)].width=4
    # adjust with for columns with areas
    col = sheet.iter_cols(min_col=col_number, max_col=col_number, min_row=2, values_only=True)
    col = sorted([len(str(x)) for y in col for x in y if x])
    if col:
        max_width = col[-1]
        sheet.column_dimensions[get_column_letter(col_number)].width=max_width


        
    


        
    



"""    
    for x in header[0]:
    if x.value == tryout[0]:
        print(x.column)

sheet["A1"] = "Jänner"
sheet["B1"] = "deivoda"
sheet["C1"] = "bamoida"

data = [["Alice", 25, "Nwq sodif"], ["Bob", 26, "lsdkjfl"], ["Charlie", 234, "csdf"]]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, cell_value in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=cell_value)
listi=[]
for index, row in enumerate(header):
    row[0].value = monthnames[index] 

tryout = ['Februar 26', 21]
"""

wb.save("excel.xlsx")
